# -*- codeing = utf-8 -*-
from openpyxl import Workbook
from openpyxl import load_workbook
import json
import os
import sys  # 导入sys模块，特别是为了使用sys.stdout.write和sys.stdout.flush来直接写入和刷新标准输出

from test_req import askURL,dowmloadPdf
from test_pdf import search_text

# 全局变量
rootPath = os.getcwd()
baseUrl = "https://www.neeq.com.cn"
search_term = "机箱" # 关键词

# 存储excel
savepath = "年度报告.xlsx"
row = 1

def main():
    # 1.创建excel表格 存储数据
    createWorkBook()

    # 2.爬取接口数据
    # 分页爬取数据
    page = 0
    for i in range(3):
        datalist = getData(page)
        # 3.保存数据
        saveData(datalist)
        page += 1
        print(f"\n完成第{page}页 !")  # 当循环完成后，打印“完成!”信息，并自动换行

# 爬取网页
def getData(page):
    global rootPath,baseUrl,search_term

    apiUrl = "https://www.neeq.com.cn/disclosureInfoController/companyAnnouncement.do?callback=jQuery331_172897375"  # 要爬取的网页链接
    payload = ("noticeType%5B%5D=5"
               "&disclosureType%5B%5D=5"
               "&disclosureSubtype%5B%5D="
               "&page=" + str(page) +
               "&companyCd="
               "&isNewThree=1"
               "&keyword=2023%E5%B9%B4%E5%B9%B4%E5%BA%A6%E6%8A%A5%E5%91%8A"
               "&xxfcbj%5B%5D=3"
               "&hyType%5B%5D="
               "&needFields%5B%5D=companyCd"
               "&needFields%5B%5D=companyName"
               "&needFields%5B%5D=disclosureTitle"
               "&needFields%5B%5D=disclosurePostTitle"
               "&needFields%5B%5D=destFilePath"
               "&needFields%5B%5D=publishDate"
               "&needFields%5B%5D=xxfcbj"
               "&needFields%5B%5D=destFilePath"
               "&needFields%5B%5D=fileExt"
               "&needFields%5B%5D=xxzrlx"
               "&siteId=1"
               "&sortfield=xxssdq"
               "&sorttype=asc")

    interfaceData = askURL(apiUrl,payload)  # 请求api 获取接口数据

    datalist = []  # 用来存储信息

    # 接口数据格式: jQuery331_1728973009099(list)
    iData_format = interfaceData[20:-1]

    iData_format_json = json.loads(iData_format)
    dataSize = len(iData_format_json)

    if dataSize > 0 :
        content_data = iData_format_json[0]["listInfo"]["content"]
        numberOfElements = iData_format_json[0]["listInfo"]["numberOfElements"]
        size = iData_format_json[0]["listInfo"]["size"]
        totalPages = iData_format_json[0]["listInfo"]["totalPages"]
        curElement = 1
        if numberOfElements >0 :
            for content in content_data:

                destFilePath = content["destFilePath"]
                destFilePathArry = destFilePath.split("/")
                pdfName = destFilePathArry[len(destFilePathArry) - 1]

                # 1.下载pdf
                pdfUrl = baseUrl + destFilePath
                dowmloadPdf(pdfUrl,rootPath,pdfName)

                # 2.识别pdf内容
                oPdfPath = rootPath + '/pdf/' + pdfName
                search_results = search_text(oPdfPath, search_term)
                # 如果文本中存在轧辊则进行保存
                if len(search_results) > 0 :
                    data = []  # 保存企业所有信息
                    data.append(content["xxfcbj"])  # 1创/0基
                    data.append(content["companyCd"])  # 代码
                    data.append(content["companyName"])  # 简称
                    data.append(content["disclosureTitle"])  # 标题
                    data.append(content["publishDate"])  # 日期
                    data.append(content["destFilePath"])  # pdf 地址
                    data.append(pdfName)  # pdf 地址
                    datalist.append(data)

                    # print(content)
                else:
                    if os.path.exists(oPdfPath):  # Delete the file
                        os.remove(oPdfPath)
                        # print(f"{oPdfPath} 不包含 {search_term} ,已删除!")
                    else:
                        print(f"{oPdfPath} does not exist.")

                simple_progress_bar(numberOfElements, curElement, page + 1, totalPages)  # 调用simple_progress_bar函数，传入总任务量和当前进度
                curElement +=1

    return datalist

# 创建表格
def createWorkBook():
    # 创建一个工作簿对象
    wb = Workbook()
    # 在索引为0的位置创建一个名为mySheet的sheet页
    ws = wb.create_sheet('年度报告', 0)
    # 对sheet页设置一个颜色（16位的RGB颜色）
    ws.sheet_properties.tabColor = 'ff72BA'

    global row
    global savepath

    col = ("类型", "代码", "简称", "标题", "日期", "pdf地址")
    for i in range(1, 7):
        ws.cell(row, i, col[i-1])  # 列名
    row +=1
    # 将创建的工作簿保存
    wb.save(savepath)

# 保存数据到表格
def saveData(datalist):

    if len(datalist) > 0:
        # print(datalist)

        global savepath
        global row

        # 加载工作簿
        wb = load_workbook(savepath)
        # 获取sheet页
        ws = wb['年度报告']

        for data in datalist:
            for i in range(1, 7):
                ws.cell(row, i, data[i-1])  # 列
            row += 1

        wb.save(savepath)


def simple_progress_bar(total, progress, page, totalPage):
    """
    显示一个简单的进度条

    :param total: 进度条的总长度（或总任务量）
    :param progress: 当前进度（已完成的任务量）
    """
    bar_length = 50  # 进度条的长度，这里设置为50个字符
    filled_length = int(round(bar_length * progress / float(total)))  # 计算已完成的进度条长度
    # 注意这里使用了float(total)确保除法结果是浮点数，然后通过round和int转换得到整数长度
    percents = round(100.0 * progress / float(total), 1)  # 计算并格式化当前进度的百分比，保留一位小数
    bar = '=' * filled_length + '-' * (bar_length - filled_length)  # 根据已完成和未完成的长度生成进度条字符串
    # 使用'='表示已完成的进度，'-'表示未完成的进度
    sys.stdout.write(f'\r进度：[{bar}] {percents}%   {page}/{totalPage}页')  # 将进度条信息写回标准输出，\r使光标回到行首
    # 这样新的进度信息就会覆盖旧的进度信息，实现进度条的更新效果
    sys.stdout.flush()  # 刷新标准输出缓冲区，确保进度条信息立即显示


if __name__ == "__main__":  # 当程序执行时
    # 调用函数
    main()
    print("爬取完毕！")
